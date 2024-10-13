import csv
import zipfile

from openpyxl.reader.excel import load_workbook
from pypdf import PdfReader

from conftest import FILE


def test_pdf_file():
    with zipfile.ZipFile(f"{FILE}/archive.zip") as zip_file:
        with zip_file.open("test_pdf.pdf") as pdf_file:
            reader = PdfReader(pdf_file)
            page = reader.pages[2]
            text = page.extract_text().strip()
            assert "The Pragmatic Bookshelf" in text


def test_csv_file():
    with zipfile.ZipFile(f"{FILE}/archive.zip") as zip_file:
        with zip_file.open("test_csv.csv") as csv_file:
            content = csv_file.read().decode('utf-8-sig')
            csvreader = list(csv.reader(content.splitlines()))
            row = csvreader[1]
            assert row[0] == "Tset"


def test_xlsx_file():
    with zipfile.ZipFile(f"{FILE}/archive.zip") as zip_file:
        with zip_file.open("test_xlsx.xlsx") as xlsx_file:
            workbook = load_workbook(xlsx_file)
            sheet = workbook.active
            assert sheet.cell(row=11, column=8).value == 5486
